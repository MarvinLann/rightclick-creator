#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel格式整理工具
功能：将Excel文件整理为标准格式
- 字体：微软雅黑 10px（表头加粗）
- 表头：灰色背景 F2F2F2 + 微软雅黑 10px 加粗 + 居中
- 数据：微软雅黑 10px，删除emoji，句号间有空行
- 对齐：内容<20字符居中，>=20字符左对齐+顶端对齐+自动换行
- 边框：全部 thin 黑色边框
- 列宽：按内容自动计算（最小10，最大60）
- 页面：横向A4，水平居中，冻结首行
"""

import re
import os
import emoji
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


def remove_emoji(text):
    """删除emoji并清理空格"""
    if not isinstance(text, str):
        return text
    text = emoji.replace_emoji(text, '')
    return text.strip()


def split_sentences_with_space(text):
    """在句号之间添加空行"""
    if not isinstance(text, str):
        return text

    sentences = re.split(r'([。！？])', text)
    result = []

    for i in range(0, len(sentences) - 1, 2):
        if sentences[i].strip():
            sentence = sentences[i].strip() + (sentences[i + 1] if i + 1 < len(sentences) else '')
            result.append(sentence)

    if len(sentences) % 2 == 1 and sentences[-1].strip():
        result.append(sentences[-1].strip())

    return '\n\n'.join(result) if result else text


def is_numeric(value):
    """检查字符串是否为数字"""
    if not value or not isinstance(value, str):
        return False
    cleaned = value.strip().replace(',', '').replace('%', '').replace('$', '').replace('¥', '')
    try:
        float(cleaned)
        return True
    except ValueError:
        return False


def is_date(value):
    """检查字符串是否为日期"""
    if not value or not isinstance(value, str):
        return False
    date_patterns = [
        r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$',
        r'^\d{1,2}[-/]\d{1,2}[-/]\d{4}$',
        r'^\d{4}年\d{1,2}月\d{1,2}日$',
        r'^\d{1,2}/\d{1,2}/\d{2,4}$',
    ]
    for pattern in date_patterns:
        if re.match(pattern, value.strip()):
            return True
    return False


def is_header_row(ws, row):
    """检查是否是表头行（关键词匹配）"""
    header_keywords = ['序号', '名称', '项目', '内容', '描述', '证明', '页码', '材料', '来源', '形式', '时间', '目的', '等级', '位置', '问题', '风险', '建议', '修改', '方向', '编号', '类型', '状态', '金额', '数量', '价格', '单价', 'qty', 'total', 'sum', 'id', 'name', 'date', 'time', 'amount', 'type', 'code', 'no.', '总计', '合计']
    keyword_count = 0
    for col in range(1, min(ws.max_column + 1, 10)):
        cell_value = str(ws.cell(row, col).value) if ws.cell(row, col).value else ''
        for keyword in header_keywords:
            if keyword in cell_value and len(cell_value) < 20:
                keyword_count += 1
                break
    return keyword_count >= 2


def detect_header_by_data_pattern(ws):
    """
    通过数据模式检测第一行是否为表头（与 CSV 转 Excel 的 detect_header 逻辑一致）
    规则1: 第一行全是字符串，第二行混合了数字/日期
    规则2: 第一行平均长度明显短于第二行（且第一行全是文本）
    """
    if ws.max_row < 2:
        return False, 0.0

    first_row = []
    second_row = []
    for col in range(1, ws.max_column + 1):
        v1 = str(ws.cell(1, col).value) if ws.cell(1, col).value is not None else ''
        v2 = str(ws.cell(2, col).value) if ws.cell(2, col).value is not None else ''
        first_row.append(v1)
        second_row.append(v2)

    if not first_row or not second_row:
        return False, 0.0

    # 规则1: 第一行全是字符串，第二行混合了数字/日期
    first_row_all_text = all(not is_numeric(cell) and not is_date(cell)
                             for cell in first_row if cell.strip())
    second_row_has_numbers = any(is_numeric(cell) for cell in second_row if cell.strip())
    second_row_has_dates = any(is_date(cell) for cell in second_row if cell.strip())

    if first_row_all_text and (second_row_has_numbers or second_row_has_dates):
        return True, 0.9

    # 规则2: 第一行平均长度明显短于数据行
    non_empty_first = [cell for cell in first_row if cell.strip()]
    non_empty_second = [cell for cell in second_row if cell.strip()]

    avg_first_len = sum(len(cell) for cell in non_empty_first) / len(non_empty_first) if non_empty_first else 0
    avg_second_len = sum(len(cell) for cell in non_empty_second) / len(non_empty_second) if non_empty_second else 0

    if avg_first_len < avg_second_len * 0.5 and first_row_all_text:
        return True, 0.7

    return False, 0.5


def identify_row_type(ws, row):
    """智能识别行类型"""
    # 优先检查是否为表头（关键词匹配）
    if is_header_row(ws, row):
        return 'header'

    # 检查第一行是否为标题（合并单元格跨多列）
    if row == 1:
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.min_col == 1:
                if merged_range.max_col - merged_range.min_col >= 2:
                    return 'title'
        # 第一行：关键词匹配失败时，用数据模式检测
        is_header, confidence = detect_header_by_data_pattern(ws)
        if is_header:
            return 'header'
        # 如果第一行包含表头关键词（放宽条件）
        header_keywords = ['序号', '名称', '项目', '内容', '描述', '证明', '页码', '材料', '来源', '形式', '时间', '目的', '等级', '位置', '问题', '风险', '建议', '修改']
        first_row_text = ' '.join(str(ws.cell(row, col).value) for col in range(1, min(ws.max_column + 1, 10)))
        if any(keyword in first_row_text for keyword in header_keywords):
            return 'header'

    if row <= 5:
        subtitle_keywords = ['致：', '案号：', '提交人：', '提交日期', '生成日期']
        cell_value = str(ws.cell(row, 1).value) if ws.cell(row, 1).value else ''
        if any(keyword in cell_value for keyword in subtitle_keywords):
            return 'subtitle'

    return 'data'


def calculate_column_widths(ws, min_row, max_row, min_col, max_col):
    """计算最佳列宽（中文字符按2个宽度计算）"""
    column_widths = {}
    for col_idx in range(min_col, max_col + 1):
        max_length = 0
        for row_idx in range(min_row, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                text = str(cell.value)
                length = sum(2 if ord(char) > 127 else 1 for char in text)
                max_length = max(max_length, min(length, 100))
        column_widths[col_idx] = min(max(max_length + 2, 10), 60)
    return column_widths


# 统一边框样式
THIN_BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color="000000")
CELL_FONT = Font(name='微软雅黑', size=10)


def apply_format(ws, row, row_type):
    """应用格式"""
    if row_type == 'title':
        font = Font(name='微软雅黑', size=14, bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = font
            cell.alignment = alignment
            cell.border = THIN_BORDER

    elif row_type == 'subtitle':
        font = Font(name='微软雅黑', size=10)
        alignment = Alignment(horizontal='left', vertical='center')
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = font
            cell.alignment = alignment
            cell.border = THIN_BORDER

    elif row_type == 'header':
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

    elif row_type == 'data':
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                cell.value = remove_emoji(cell.value)
                if col >= 3 and '。' in cell.value:
                    cell.value = split_sentences_with_space(cell.value)

                text_length = len(cell.value.strip())
                if text_length < 20:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = CELL_FONT
            cell.border = THIN_BORDER


def format_excel(input_file, output_file):
    """格式化Excel文件"""
    # 检查文件格式，如果是XLS，先转换为XLSX
    if input_file.lower().endswith('.xls'):
        import tempfile
        df_dict = pd.read_excel(input_file, sheet_name=None)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                for sheet_name, df in df_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            tmp_path = tmp.name
        wb = load_workbook(tmp_path)
        ws = wb.active
        os.unlink(tmp_path)
    else:
        wb = load_workbook(input_file)
        ws = wb.active

    # 先识别所有行的类型，判断是否有表头
    has_header = False
    row_types = {}
    for row in range(1, ws.max_row + 1):
        row_type = identify_row_type(ws, row)
        row_types[row] = row_type
        if row_type == 'header':
            has_header = True

    # 处理每一行
    for row in range(1, ws.max_row + 1):
        apply_format(ws, row, row_types[row])

    # 自动计算列宽
    column_widths = calculate_column_widths(ws, 1, ws.max_row, 1, ws.max_column)
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # 设置行高
    for row in range(1, ws.max_row + 1):
        if row_types.get(row) == 'header':
            ws.row_dimensions[row].height = 25
        elif row_types.get(row) == 'title':
            ws.row_dimensions[row].height = 35
        else:
            ws.row_dimensions[row].height = None  # 自动

    # 页面设置：横向A4，水平居中
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=1, right=1, top=1, bottom=1)
    ws.print_options.horizontalCentered = True

    # 冻结首行（如果有表头）
    if has_header:
        ws.freeze_panes = 'A2'

    wb.save(output_file)
    print(f"已保存到: {output_file}")


if __name__ == "__main__":
    import sys
    import os
    
    if len(sys.argv) < 2:
        print("用法: python excel_format.py <excel文件路径>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    if not os.path.exists(input_file):
        print(f"错误: 文件不存在: {input_file}")
        sys.exit(1)
    
    if not input_file.lower().endswith(('.xlsx', '.xls')):
        print("错误: 只支持Excel文件")
        sys.exit(1)
    
    # 获取输入文件所在目录和文件名
    input_dir = os.path.dirname(input_file)
    input_filename = os.path.basename(input_file)
    output_filename = os.path.splitext(input_filename)[0] + '_格式化.xlsx'
    
    # 在同一目录下生成输出文件
    if input_dir:
        output_file = os.path.join(input_dir, output_filename)
    else:
        output_file = output_filename
    
    format_excel(input_file, output_file)
